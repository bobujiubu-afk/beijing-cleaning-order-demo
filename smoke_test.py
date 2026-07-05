from io import BytesIO
import re
import sqlite3
from datetime import date

from openpyxl import load_workbook

from app import DATABASE, app, init_db, seed_demo_data


def main():
    init_db()
    seed_demo_data(force=True)
    client = app.test_client()

    admin = client.get("/admin", follow_redirects=False)
    assert admin.status_code == 200 and "老板后台登录".encode("utf-8") in admin.data
    assert "默认演示密码".encode("utf-8") not in admin.data
    assert admin.headers["Cache-Control"].startswith("no-store")
    assert client.get("/api/orders").status_code == 401
    assert client.get("/api/order-counts").status_code == 401
    assert client.get("/api/push-public-key").status_code == 401

    bad = client.post("/submit", data={"customer_name": "", "phone": "", "address": ""})
    assert bad.status_code == 400

    test_names = ["测试客户自测一", "测试客户自测二", "测试客户自测三"]
    for index, test_name in enumerate(test_names, start=1):
        good = client.post(
            "/submit",
            data={
                "customer_name": test_name,
                "phone": f"1390000000{index}",
                "address": f"北京市测试区{index}号",
                "service_type": "开荒保洁",
                "area": "88㎡",
                "appointment_date": date.today().strftime("%Y-%m-%d"),
                "appointment_period": "下午",
                "need_invoice": "否",
                "source": "微信",
                "remark": "",
            },
            follow_redirects=False,
        )
        assert good.status_code == 302 and "/submit/success" in good.headers["Location"]

    login = client.post("/login", data={"password": "123456"}, follow_redirects=True)
    assert login.status_code == 200 and "老板后台".encode("utf-8") in login.data
    assert "已报价".encode("utf-8") not in login.data
    assert "已完成".encode("utf-8") not in login.data
    assert "phone-link".encode("utf-8") in login.data
    assert "new-badge".encode("utf-8") in login.data
    assert "开启页面声音".encode("utf-8") in login.data
    assert "开启手机消息推送".encode("utf-8") in login.data
    assert "手机值守模式".encode("utf-8") not in login.data
    assert "watchOverlay".encode("utf-8") not in login.data
    assert "今天".encode("utf-8") in login.data
    assert "有新的客户预约，请尽快联系".encode("utf-8") in login.data
    push_key = client.get("/api/push-public-key")
    assert push_key.status_code == 200 and len(push_key.json["publicKey"]) > 60
    push_test = client.post("/api/push-test")
    assert push_test.status_code == 200 and push_test.json["sent"] == 0

    api_orders = client.get("/api/orders")
    assert api_orders.status_code == 200
    api_data = api_orders.get_json()
    assert api_data["pending_count"] >= 3
    assert api_data["new_count"] >= 3
    assert api_data["latest_order_id"] > 0
    assert api_data["latest_order"]["customer_name"] == "测试客户自测三"
    assert "测试客户自测三" in api_data["html"]
    assert "预约日期".encode("utf-8") in login.data
    assert "订单金额".encode("utf-8") in login.data
    assert api_data["html"].find("测试客户自测三") < api_data["html"].find("测试客户自测二")
    assert "已报价" not in api_data["html"]
    assert "已完成" not in api_data["html"]

    qr_page = client.get("/appointment-qr")
    assert qr_page.status_code == 200 and "客户预约链接".encode("utf-8") in qr_page.data
    qr_png = client.get("/appointment-qr.png")
    assert qr_png.status_code == 200 and qr_png.data[:8] == b"\x89PNG\r\n\x1a\n"
    assert "预约二维码".encode("utf-8") in client.get("/admin").data

    reverse_bad = client.post("/submit/reverse-geocode", json={"latitude": "bad", "longitude": "bad"})
    assert reverse_bad.status_code == 400
    submit_page = client.get("/submit")
    assert submit_page.status_code == 200
    assert "自动定位".encode("utf-8") in submit_page.data

    conn = sqlite3.connect(DATABASE)
    order_row = conn.execute(
        "SELECT id FROM orders WHERE customer_name = ? AND deleted_at IS NULL ORDER BY id DESC",
        (test_names[-1],),
    ).fetchone()
    assert order_row is not None
    order_id = order_row[0]
    conn.close()

    update = client.post(
        f"/orders/{order_id}/update",
        json={
            "status": "已联系",
        },
    )
    assert update.status_code == 200
    amount_update = client.post(f"/orders/{order_id}/update", json={"amount": "580"})
    assert amount_update.status_code == 200
    after_contact = client.get("/api/orders").get_json()
    assert "测试客户自测三" in after_contact["html"]
    contacted_section = after_contact["html"][after_contact["html"].find("测试客户自测三") - 300:after_contact["html"].find("测试客户自测三") + 300]
    assert "new-badge" not in contacted_section

    deal = client.post(
        f"/orders/{order_id}/update",
        json={
            "status": "已成交",
        },
    )
    assert deal.status_code == 200

    export = client.get("/export", follow_redirects=True)
    assert export.status_code == 200 and "本次导出的订单明细".encode("utf-8") in export.data
    match = re.search(rb'href="([^"]*/exports/download/[^"]+)"', export.data)
    assert match
    download_url = match.group(1).decode("utf-8")
    excel = client.get(download_url)
    assert excel.status_code == 200
    workbook = load_workbook(BytesIO(excel.data))
    assert workbook.active.max_row >= 2
    headers = [cell.value for cell in workbook.active[1]]
    assert "预约日期" in headers
    assert "预约时间段" in headers
    assert "订单金额" in headers
    assert "报价金额" not in headers
    assert "成交金额" not in headers
    assert "负责人" not in headers
    assert "回访状态" not in headers
    exported_statuses = [cell.value for cell in workbook.active["J"][1:]]
    assert "已报价" not in exported_statuses
    assert "已完成" not in exported_statuses

    backup = client.get("/backup/database")
    assert backup.status_code == 200 and len(backup.data) > 1000

    delete = client.post(f"/orders/{order_id}/delete")
    assert delete.status_code == 302
    conn = sqlite3.connect(DATABASE)
    deleted_at = conn.execute("SELECT deleted_at FROM orders WHERE id = ?", (order_id,)).fetchone()[0]
    visible_count = conn.execute("SELECT COUNT(*) FROM orders WHERE deleted_at IS NULL").fetchone()[0]
    conn.close()
    assert deleted_at is not None
    assert visible_count == 7

    seed_demo_data(force=True)
    print("Smoke test passed.")


if __name__ == "__main__":
    main()
